#!/usr/bin/env python3
# -*- coding: utf-8 -*-
#
# This program is free software: you can redistribute it and/or modify
#     it under the terms of the GNU General Public License as published by
#     the Free Software Foundation; version 3 only.
#
#     This program is distributed in the hope that it will be useful,
#     but WITHOUT ANY WARRANTY; without even the implied warranty of
#     MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#     GNU General Public License for more details.
#
#     You should have received a copy of the GNU General Public License
#     along with this program.  If not, see <http://www.gnu.org/licenses/>.


"""Xparse parses xls and extracts structured data."""

import argparse
import re
import os
import logging
import json
from collections import OrderedDict
import openpyxl
from dicttoxml2 import dicttoxml2
#from string import punctuation

#logging config:
logger = logging.getLogger(__name__) #pylint: disable=invalid-name
logger.setLevel(logging.DEBUG)
fh = logging.FileHandler('xparse.log') #pylint: disable=invalid-name
ch = logging.StreamHandler() #pylint: disable=invalid-name
formatter_ch = logging.Formatter(fmt='[%(levelname)s] %(message)s') #pylint: disable=invalid-name
formatter_fh = logging.Formatter(fmt='[%(levelname)s] %(asctime)s %(message)s', datefmt='%H:%M:%S') #pylint: disable=invalid-name
ch.setFormatter(formatter_ch)
fh.setFormatter(formatter_fh)
ch.setLevel(logging.INFO)
fh.setLevel(logging.DEBUG)
logger.addHandler(ch)
logger.addHandler(fh)
#logging config.

# Some globals
# Load dictionaries for objects
with open('dictionaries.json', 'r') as file_in:
    dictionaries = json.load(file_in) #pylint: disable=invalid-name

############################
# General helper functions
############################

def get_sorted_coord(coord_lst):
    """Sort alphanum coordinates"""
    return sorted(coord_lst, key=lambda x: int(re.findall(r'\d+$', x)[0]))


def validate_dimensions(dimensions):
    """Check dimensions against pattern `AAnn:BB:nn`"""
    valid = re.compile(r'^[a-zA-Z]+[0-9]+:[a-zA-Z]+[0-9]+$')
    return bool(valid.match(dimensions))


def value_from_dict(value, dictionary='none_values'):
    """Get a dictionary values"""
    def normalize(value):
        """Normalize input value"""
        value = value.lower().strip()
        return value

    if value is None:
        return value
    try:
        return dictionaries[dictionary][normalize(value)]
    except Exception:
        logger.warning('"%s" not in <%s>', value, dictionary)
        return value


def not_empty(val):
    """Check value for empty and unwanted values"""
    unwanted = [
        'автомобиль легковой:',
        'автомобили легковые:',
        'автоприцеп:',
        'водный транспорт:',
        'мототранспортные средства:',
        'автоприцепы:',
        'иные транспортные средства:',
        'мототранспортное средство:'
        ]
    empty_values = ['-', ' ', '', 'не имеет', None]
    if type(val).__name__ in ('str', 'unicode'):
        val = " ".join(val.lower().split()) # catch ' -', ' не  имеет '...
    if type(val).__name__ in ('int', 'float'):
        val = str(val)
    return bool(val not in empty_values and val not in unwanted)



#######################################
# Collect data and parse persons
#######################################

def parse_person(column_range):
    """Parse a person from a slot"""
    try:
        start, end = column_range.split(':')
        logger.info('Parsing persons from %s to %s', start, end)
    except Exception as err:
        logger.error('Invalid column range, %s', err)

    unwanted_chars = '!"#$%&\'()*+,-./:;<=>?@[\\]^_`{|}~ '
    parsed_persons = []
    a_slots = get_slot(start, end)
    person_id = 1
    p = 0
    # a_slots = [{A2, A17, 1}, {A18, A29, 2}]
    for a_slot in a_slots:
        # a_slot = {A2, A17, 1}
        b_start = shift_col(a_slot['start'])
        b_end = shift_col(a_slot['end'])
        b_slots = get_slot(b_start, b_end)
        p_raw = str(a_slot['value'])

        try:
            p_raw_int = int(p_raw.strip(unwanted_chars))
        except ValueError as err:
            logger.warning('Strange P numbering: %s', p_raw)
            p_raw_int = p_raw.strip(unwanted_chars)

        person_num = 1
        p += 1
        if not ws[b_start].value:
            logger.warning('Person at %s <%s>.', b_start, ws[b_start].value)
        else:
            for b_slot in b_slots:
                #if b['start'] b_slot['value']
                try:
                    position_coord = shift_col(b_slot['start'])
                    position = ws[position_coord].value
                    income_coord = shift_col(position_coord)
                    income = ws[income_coord].value
                    p_start = b_slot['start']
                    p_end = b_slot['end']
                    p_name = b_slot['value']
                    parsed_persons.append({
                        'p_raw':p_raw_int,
                        'p':p,
                        'person_id':person_id,
                        'person_num':person_num,
                        'start':p_start,
                        'end': p_end,
                        'name': p_name,
                        'position': position,
                        'income': income,
                        'ownership':parse_ownership(b_slot),
                        'usage':parse_usage(b_slot),
                        'vehicle':parse_vehicle(b_slot)
                        })
                    person_id += 1
                    person_num += 1
                except Exception as err:
                    logger.error('Error while parsing persons: %s', err)

    # check whether p_old == p
    ps_generated = [parsed_persons[i]['p'] for i in range(len(parsed_persons))]
    ps_from_file = [parsed_persons[i]['p_raw'] for i in range(len(parsed_persons))]
    check_result = check_lists_mismatch(ps_generated, ps_from_file)
    if check_result:
        logger.warning('P numbering mismatch at %s', check_result)
    return parsed_persons


def check_lists_mismatch(list_a, list_b):
    """Check whether p-numbering is wrong in the file"""
    from itertools import zip_longest
    for a, b in zip_longest(list_a, list_b):
        if a == b:
            pass
        else:
            return a, b


def get_slot(start, end):
    """Given /start/ and /end/ coordinates of a range, get all slots,
       i.e. assumed range till the next slot"""
    def not_false_empty(val):
        """Check whether cell.value is not false empty, e.g. ' ' or '  -  '"""
        empty_values = ['-', ' ', '', None]
        if type(val).__name__ in ('str', 'unicode'):
            val = " ".join(val.lower().split()) # catch ' -', ' не  имеет '...
        return bool(val not in empty_values)

    data = []
    data_cache = {}
    for row in ws[start:end]:
        for cell in row:
            if not_false_empty(cell.value):
                if data_cache:
                    data.append(data_cache)
                    data_cache = {}
                    data_cache['start'] = cell.coordinate
                    data_cache['value'] = cell.value
                    data_cache['end'] = cell.coordinate
                    #continue
                    break
                else:
                    data_cache['start'] = cell.coordinate
                    data_cache['value'] = cell.value
                    data_cache['end'] = cell.coordinate
                    #continue
                    break
            else:
                if data_cache:
                    data_cache['end'] = cell.coordinate
                else:
                    #continue
                    break
    # check for remaining in cache
    if data_cache:
        data.append(data_cache)
        data_cache = {}
    return data


def shift_col(col, step=1):
    """Return shifter column index
       FIXME: Support for AA, GB indeces"""
    a_z = range(ord('A'), ord('Z') +1)
    col_letter_ordinal = ord(col[:1]) + step
    col_num = col[1:]
    if col_letter_ordinal in a_z:
        shifted_col = chr(col_letter_ordinal) + col_num
        return shifted_col
    elif col_letter_ordinal > 90:
        logger.error("Can't shift '%s' behind 'Z%s'", col, col_num)
    else:
        logger.error('Column "%s" not in A-Z range', col)


def parse_ownership(person_slot):
    """person_slot is dict"""
    ownership_list = []
    start = shift_col(person_slot['start'], 3)
    end = shift_col(person_slot['end'], 3)
    for row in ws[start:end]:
        for cell in row:
            coord = cell.coordinate
            if cell.value:
                ownership_list.append({
                    'own_obj':cell.value,
                    'own_type':ws[shift_col(coord, 1)].value,
                    'own_sq':ws[shift_col(coord, 2)].value,
                    'own_location':ws[shift_col(coord, 3)].value
                    })
            elif not cell.value and ws[shift_col(coord)].value not in ['-', None]:
                # checking whether cell to the right is not empty
                logger.warning('Value missing: %s?', coord)
                ownership_list.append({
                    'own_obj': 'иное', # cell.value to dafult
                    'own_type':ws[shift_col(coord, 1)].value,
                    'own_sq':ws[shift_col(coord, 2)].value,
                    'own_location':ws[shift_col(coord, 3)].value
                    })
            else:
                #logger.debug('Line "{0}" empty'.format(coord[1:]))
                pass
    return ownership_list


def parse_usage(person_slot):
    """Parse 'use_*' columns"""
    usage_list = []
    start = shift_col(person_slot['start'], 7)
    end = shift_col(person_slot['end'], 7)
    for row in ws[start:end]:
        for cell in row:
            coord = cell.coordinate
            if cell.value:
                usage_list.append({
                    'use_obj':cell.value,
                    'use_sq': ws[shift_col(coord, 1)].value,
                    'use_loc':ws[shift_col(coord, 2)].value
                    })
            elif not cell.value and ws[shift_col(coord)].value not in ['-', None]:
                logger.info('Value missing: %s?', coord)
                usage_list.append({
                    'use_obj': 'иное',#cell.value,
                    'use_sq': ws[shift_col(coord, 1)].value,
                    'use_loc':ws[shift_col(coord, 2)].value
                    })
            else:
                pass
                #logger.debug('Line "{0}" empty'.format(coord[1:]))
    return usage_list


def parse_vehicle(person_slot):
    """Parse 'vehicle_*' columns"""
    vehicle_list = []
    start = shift_col(person_slot['start'], 10)
    end = shift_col(person_slot['end'], 10)
    for row in ws[start:end]:
        for cell in row:
            coord = cell.coordinate
            if cell.value:
                vehicle_list.append({
                    'vehicle_item':cell.value,
                    'vehicle_pay':ws[shift_col(coord, 1)].value
                    })
            elif not cell.value and ws[shift_col(coord)].value not in ['-', None]:
                logger.warning('Value missing at %s?', coord)
            else:
                pass
    return vehicle_list


################################
# Modify collected data
################################

def map_data(person_data):
    """Transfers/maps data from a dict to an OrderedDict"""
    name = set_name(person_data)
    #relationType = None
    position = set_position(person_data)
    realties = []
    for realty in person_data['ownership']: # in ownership
        if not_empty(realty['own_obj']):
            own_type, own_part = set_ownership(realty)
            realty_data = OrderedDict()
            #realty_data['realtyType_'] = 'В собственности'
            realty_data['realtyType'] = '1'
            #realty_data['objectType_'] = realty['own_obj']
            realty_data['objectType'] = value_from_dict(realty['own_obj'], 'objectType')
            #realty_data['ownershipType_'] = own_type
            realty_data['ownershipType'] = value_from_dict(own_type, 'ownershipType')
            realty_data['ownershipPart'] = own_part
            realty_data['square'] = realty['own_sq']
            #realty_data['country_num'] = realty['own_location']
            realty_data['country'] = value_from_dict(realty['own_location'], 'country')
            realties.append(realty_data)
        else:
            logger.debug('OWN_OBJ EMPTY: %s', realty['own_obj'])
    for realty in person_data['usage']: # in use
        if not_empty(realty['use_obj']):
            realty_data = OrderedDict()
            realty_data['realtyType'] = '2' # in use:2
            #realty_data['realtyType_'] = 'В пользовании'
            realty_data['objectType'] = value_from_dict(realty['use_obj'], 'objectType')
            realty_data['square'] = realty['use_sq']
            #realty_data['country'] = realty['use_loc']
            realty_data['country'] = value_from_dict(realty['use_loc'], 'country')
            realties.append(realty_data)
        else:
            logger.debug('"USE_OBJ EMPTY: %s', realty['use_obj'])

    transports = []

    for transport in person_data['vehicle']:
        if not_empty(transport['vehicle_item']):
            transports.append(OrderedDict({
                'transportName':transport['vehicle_item']
                }))
        else:
            logger.debug('TRANSPORT EMPTY: %s', transport['vehicle_item'])
    income = set_income(person_data)
    income_comment = None # disabled
    income_source = None # disabled

    pers = OrderedDict()
    pers['id'] = person_data['person_id']
    #pers['p'] = person_data['p']
    #pers['p_raw'] = person_data['p_raw']
    pers['name'] = name
    pers['relativeOf'] = person_data['relativeOf']
    pers['relationType'] = value_from_dict(person_data['relationType'], 'relationType')
    pers['position'] = position

    if realties:
        pers['realties'] = realties
    else:
        pers['realties'] = None

    if transports:
        pers['transports'] = transports
    else:
        pers['transports'] = None

    pers['income'] = income
    pers['incomeComment'] = income_comment
    pers['incomeSource'] = income_source
    return pers


def set_name(person_data):
    """If relativeOf, set name=None, else name"""
    types_of_relatives = ['супруг',
                          'супруга',
                          'несовершеннолетний ребенок',
                          'несовершеннолетний ребёнок']
    if person_data['relativeOf']:
        return None
    elif person_data['name'] in types_of_relatives:
        logger.warning('Missing person: P=%s at %s', person_data['p'], person_data['start'])
    else:
        return person_data['name']


def set_position(person_data):
    """Set position value"""
    if not person_data['relativeOf']:
        return person_data['position']


def get_ownpart_amount(own_type_str):
    """Get the amount of ownershipPart"""
    pattern = r'[0-9]+\s?[,/.]\s?[0-9]+|[0-9]+'
    found = re.search(pattern, own_type_str)
    if found:
        amount = found.group(0)
        return amount.replace(' ', '') # ugly duck
    else:
        return own_type_str


def set_ownership(realty):
    """Set ownershipType, ownershipPart
    FIXME: not all patterns given OR reimplement"""
    own_type = realty['own_type']
    own_part = None
    try:
        if re.search(r'дол', own_type.lower()):
            logger.debug('match: %s', own_type)
            own_part = get_ownpart_amount(realty['own_type'])
            own_type = 'долевая'

        elif re.search(r'инди', own_type.lower()):
            logger.debug('match: %s', own_type)
            own_type = 'индивидуальная'

        elif re.search(r'местн', own_type.lower()):
            logger.debug('match: %s', own_type)
            own_type = 'совместная'

        else:
            logger.info('Ownership unknown: %s', own_type)
    except TypeError:
        logger.warning('Ownership invalid <%s>.', own_type)
    finally:
        return own_type, own_part


def set_income(person_data):
    """Set income for person"""
    if not_empty(person_data['income']):
        return person_data['income']


######################################
# Prepare data before saving to xml
######################################

def get_p(all_data):
    """Get respective p"""
    last = all_data[-1]['person_id'] # 1-99
    blocks = []
    for i in range(last):
        b = all_data[i]['p']
        if b not in blocks:
            blocks.append(b)
    return blocks


def make_blocks(data):
    """Divide into blocks according to `main person`"""
    blocks = []
    for num in get_p(data):
        block = [b for b in data if b['p'] == num]
        blocks.append(block)
    return blocks


def set_relations(blocks_by_person):
    """Add relationship information to every person in every block"""
    for block in blocks_by_person:
        main = block[0]['person_id']
        for person in block:
            if person['person_num'] == 1:
                person['relativeOf'] = None
                person['relationType'] = None
            else:
                person['relativeOf'] = main
                person['relationType'] = person['name']



##########################
# Load and save data
##########################

def load_file(xls_file):
    """Loading file"""
    try:
        logger.info('Loading data from %s...', xls_file)
        workbook = openpyxl.load_workbook(xls_file, read_only=False)
        worksheet_name = workbook.sheetnames[0]
        worksheet = workbook[worksheet_name]
        logger.info('Data loaded.')
    except Exception as err:
        logger.error('Error (%s) loading file: %s', err, xls_file)
    return worksheet


def parent_to_child(parent_name):
    """Change xml item name to be singular of its parent item"""
    parents = {'realties':'realty',
               'transports':'transport',
               'persons':'person'}
    if parent_name in parents.keys():
        return parents[parent_name]


def save_to_file(blocks_of_data, split_at=0, save_dir='out'):
    """"Iterate over a list of blocks with common 'p' and save to .xml
    FIXME: add leading zeros to file names"""
    def save(input_data, sdir, output_xml):
        """xml -> file.xml"""
        xml_data = dicttoxml2.dict2xml(input_data,
                                       attr_type=False,
                                       item_func=parent_to_child,
                                       custom_root='persons')
        with open(sdir + os.sep + output_xml, 'wb') as f:
            f.write(xml_data)

    #if type(split_at) != type(2):
    if not isinstance(split_at, int):
        split_at = 0
    if split_at < 0:
        split_at = 0

    try:
        os.mkdir(save_dir)
    except OSError as exc:
        if exc.errno == 17:
            logger.info('Directory "%s" already exists', save_dir)
    except Exception as err:
        logger.error('%s. Couldn\'t create directory', err)

    blocks_count = 0
    persons_count = 0
    persons_list = []

    for block in blocks_of_data:
        blocks_count += 1
        for person in block:
            p = map_data(person)
            persons_list.append(p)
            persons_count += 1

        if split_at > 0 and blocks_count % split_at == 0:
            file_num = str(blocks_count + 1 - split_at) + '-' + str(blocks_count)
            file_name = 'persons-' + file_num + '.xml'
            save(persons_list, save_dir, file_name)
            persons_list = []

    if persons_list:
        if split_at > 0:
            file_num = str(blocks_count - (blocks_count % split_at) + 1) + '-' + str(blocks_count)
            file_name = 'persons-' + file_num + '.xml'
        if split_at == 0:
            file_num = str(blocks_count)
            file_name = 'persons-' + file_num + '.xml'
        save(persons_list, save_dir, file_name)
    logger.info('Total blocks in XML: %s / persons: %s.', blocks_count, persons_count)



if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    """
    Example: xparse.py input.xlsx -c A2:A1000 -s 20
    """
    parser.add_argument("xls_file",
                        help="Input xls file")
    parser.add_argument("-c", "--column_range",
                        help="Persons column range - parse persons from Ax to Axxx",
                        type=str)
    parser.add_argument("-s", "--split_at",
                        help="Split output xml file into N blocks each",
                        type=int,
                        default=0)
    parser.add_argument("-t", "--save_dir",
                        help="Directory to save files to",
                        type=str, default='out')
    ARGS = parser.parse_args()

    #test_get_slot('C2', 'C9')
    #test_parse_person('A2', 'A39')
    #ws = load_file('data/book_101.xlsx')
    #data_all = parse_person('A2:A787')

    if validate_dimensions(ARGS.column_range):
        logger.info('Dimensions valid.')
        try:
            ws = load_file(ARGS.xls_file)
            data_all = parse_person(ARGS.column_range)
            blocks_by_p = make_blocks(data_all)
            set_relations(blocks_by_p)
            save_to_file(blocks_by_p, ARGS.split_at, ARGS.save_dir)
        except Exception as err:
            logger.error("Something is wrong. %s", err)
    else:
        logger.error('Wrong dimensions')
